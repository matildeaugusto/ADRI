import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from numpy.random import randint


# consumer counts to test
consumer_counts = [5, 15, 25, 35, 45]

ts = 20
te = 96
N = 50  # repetitions per epsilon

noise_levels = [0.0, 0.01, 0.02, 0.05, 0.1, 0.12, 0.15, 0.2, 0.25, 0.3, 0.5]

# READ EXCEL DATA

wb = load_workbook('Prob1_Conso_Data.xlsx')
ws = wb.active

raw_list = []
for row in ws.iter_rows(values_only=True):
    raw_list.append(list(row))

raw_data = np.array(raw_list)

checks = 0
nr = 1
data = np.zeros((1, 96))
h = raw_data[0:96, 0]

for i in range(1, raw_data.shape[0] + 1):
    if raw_data[i - 1, 0] == h[checks]:
        checks += 1
    else:
        checks = 0

    if checks == 96:
        if np.sum(raw_data[i - 96:i, 1]) != 0:
            data[nr - 1, :] = raw_data[i - 96:i, 1]
            data = np.vstack([data, np.zeros((1, 96))])
            nr += 1
        checks = 0

data = data[:-1, :]

# compute_accuracy USING SVD

def compute_accuracy(consumptions, phases_true, ts, te, noise):
    pw = consumptions[:, ts - 1:te]
    power = 4 * pw
    A = power.T
    n_periods, nc = A.shape

    phase_idx_true = phases_true - 1

    Y_true = np.zeros((n_periods, 3))
    for f in range(3):
        Y_true[:, f] = A[:, phase_idx_true == f].sum(axis=1)

    Y_noisy = Y_true * (1 + noise * np.random.randn(*Y_true.shape))

    U, S, Vt = np.linalg.svd(A, full_matrices=False)
    S_inv = np.diag(1 / S)

    B = np.zeros((nc, 3))
    for f in range(3):
        b = Y_noisy[:, f]
        x = Vt.T @ S_inv @ U.T @ b
        x[x < 0] = 0
        B[:, f] = x

    phases_est = np.argmax(B, axis=1) + 1
    accuracy = np.mean(phases_est == phases_true)

    return accuracy, phases_est, B

# FIND CRITICAL ε FOR ONE SCENARIO

def epsilon_critico(original, phase, noise):
    epsilons = np.concatenate([
        np.logspace(-5, -1, 40),
        np.linspace(0.1, 2, 40)
    ])

    for eps in epsilons:
        count = 0

        for _ in range(N):
            data_test = original.copy()
            data_test[1, :] = original[1, :] * (1 + eps)

            accuracy, phases_est, _ = compute_accuracy(data_test, phase, ts, te, noise)

            if phases_est[1] == phase[1]:
                count += 1

        if count / N >= 0.95:
            return eps

    return None

# COMPUTE CRITICAL ε FOR EACH CONSUMER COUNT


eps_matrix = np.zeros((len(consumer_counts), len(noise_levels)))

for i, nc in enumerate(consumer_counts):
    print(f"\n=== Testing nc = {nc} consumers ===")

    original_nc = data[0:nc, :]
    phase_nc = randint(1, 4, nc)

    for j, noise in enumerate(noise_levels):
        eps_crit = epsilon_critico(original_nc, phase_nc, noise)
        eps_matrix[i, j] = eps_crit
        print(f"Noise {noise:.3f} → critical ε = {eps_crit}")

plt.figure(figsize=(10,6))

for i, nc in enumerate(consumer_counts):
    plt.plot(noise_levels, eps_matrix[i], '-o', label=f"{nc} consumers")

plt.xlabel("Noise level")
plt.ylabel("Critical ε")
plt.title("Critical ε vs noise for different consumer counts")
plt.grid(True)
plt.legend()
plt.show()
